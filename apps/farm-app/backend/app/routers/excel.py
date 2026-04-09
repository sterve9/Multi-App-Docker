from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import extract
from io import BytesIO
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .. import models
from ..database import get_db
from ..auth import get_current_user

router = APIRouter(prefix="/excel", tags=["excel"])

# ── Couleurs ──────────────────────────────────────────────────────────
GREEN_DARK  = "1a5c38"
GREEN_MED   = "2d7a4f"
GREEN_LIGHT = "e8f5ee"
GREY_LIGHT  = "f5f5f5"
WHITE       = "ffffff"
RED_LIGHT   = "fde8e8"
AMBER_LIGHT = "fff8e1"


def _header_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _thin_border() -> Border:
    s = Side(style="thin", color="dddddd")
    return Border(left=s, right=s, top=s, bottom=s)


def _set_col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


def _title_row(ws, text: str, nb_cols: int, row: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nb_cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, color=WHITE, size=12)
    cell.fill = _header_fill(GREEN_DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22


def _col_headers(ws, headers: list, row: int):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = _header_fill(GREEN_MED)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
    ws.row_dimensions[row].height = 18


def _data_row(ws, values: list, row: int, alternate: bool = False):
    fill = _header_fill(GREEN_LIGHT) if alternate else _header_fill(WHITE)
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = fill
        cell.border = _thin_border()
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 16


# ── Feuille Résumé ────────────────────────────────────────────────────
def _sheet_resume(wb, ferme, annee, parcelles, recoltes, mouvements, recommandations):
    ws = wb.active
    ws.title = "Résumé"

    _title_row(ws, f"Farm Manager — {ferme.nom} — Bilan {annee}", 4, 1)
    ws.row_dimensions[2].height = 8

    total_kg = sum(r.quantite_kg for r in recoltes)
    total_valeur = sum(r.quantite_kg * (r.prix_kg or 0) for r in recoltes)
    total_couts = sum(m.quantite * (m.cout_unitaire or 0) for m in mouvements if m.type_mouvement == models.TypeMouvementEnum.entree)
    marge = total_valeur - total_couts

    kpis = [
        ("Ferme", ferme.nom),
        ("Localisation", ferme.localisation or "—"),
        ("Surface (ha)", ferme.surface_ha or "—"),
        ("Année", annee),
        ("Nombre de parcelles", len(parcelles)),
        ("Total récoltes (kg)", round(total_kg, 1)),
        ("Valeur récoltes (TND)", round(total_valeur, 2)),
        ("Total dépenses (TND)", round(total_couts, 2)),
        ("Marge brute (TND)", round(marge, 2)),
        ("Recommandations actives", sum(1 for r in recommandations if r.statut == models.StatutRecommandationEnum.en_attente)),
    ]

    for i, (label, value) in enumerate(kpis):
        row = i + 3
        cell_l = ws.cell(row=row, column=1, value=label)
        cell_l.font = Font(bold=True, size=10)
        cell_l.fill = _header_fill(GREY_LIGHT)
        cell_l.border = _thin_border()
        cell_l.alignment = Alignment(vertical="center")

        cell_v = ws.cell(row=row, column=2, value=value)
        cell_v.border = _thin_border()
        cell_v.alignment = Alignment(vertical="center")
        if label == "Marge brute (TND)":
            cell_v.font = Font(bold=True, color="1a5c38" if marge >= 0 else "c0392b", size=11)

        ws.row_dimensions[row].height = 16

    _set_col_width(ws, 1, 30)
    _set_col_width(ws, 2, 22)


# ── Feuille Récoltes ──────────────────────────────────────────────────
def _sheet_recoltes(wb, recoltes, parcelles_map, annee):
    ws = wb.create_sheet("Récoltes")
    headers = ["Date", "Parcelle", "Variété", "Quantité (kg)", "Qualité", "Prix/kg (TND)", "Valeur (TND)", "Destination", "Notes"]
    _title_row(ws, f"Récoltes {annee}", len(headers), 1)
    _col_headers(ws, headers, 2)

    for i, r in enumerate(recoltes):
        p = parcelles_map.get(r.parcelle_id)
        valeur = round(r.quantite_kg * (r.prix_kg or 0), 2)
        _data_row(ws, [
            r.date.strftime("%d/%m/%Y") if r.date else "",
            p.nom if p else "—",
            p.variete.value if p else "—",
            round(r.quantite_kg, 1),
            r.qualite.value if r.qualite else "—",
            r.prix_kg or 0,
            valeur,
            r.destination or "—",
            r.notes or "",
        ], row=i + 3, alternate=i % 2 == 0)

    # Ligne total
    total_row = len(recoltes) + 3
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=round(sum(r.quantite_kg for r in recoltes), 1)).font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=round(sum(r.quantite_kg * (r.prix_kg or 0) for r in recoltes), 2)).font = Font(bold=True)
    for c in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=c).fill = _header_fill(GREEN_LIGHT)
        ws.cell(row=total_row, column=c).border = _thin_border()

    widths = [14, 18, 14, 14, 16, 14, 14, 18, 20]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)


# ── Feuille Traitements ───────────────────────────────────────────────
def _sheet_traitements(wb, traitements, parcelles_map, annee):
    ws = wb.create_sheet("Traitements")
    headers = ["Date", "Parcelle", "Type", "Produit", "Dose", "Unité", "Notes"]
    _title_row(ws, f"Traitements {annee}", len(headers), 1)
    _col_headers(ws, headers, 2)

    for i, t in enumerate(traitements):
        p = parcelles_map.get(t.parcelle_id)
        _data_row(ws, [
            t.date.strftime("%d/%m/%Y") if t.date else "",
            p.nom if p else "—",
            t.type_traitement.value if t.type_traitement else "—",
            t.produit or "—",
            t.dose or "—",
            t.unite or "—",
            t.notes or "",
        ], row=i + 3, alternate=i % 2 == 0)

    widths = [14, 18, 14, 22, 10, 10, 24]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)


# ── Feuille Stocks ────────────────────────────────────────────────────
def _sheet_stocks(wb, stocks):
    ws = wb.create_sheet("Stocks")
    headers = ["Produit", "Catégorie", "Quantité", "Unité", "Seuil alerte", "Coût unitaire (TND)", "Valeur stock (TND)", "Notes"]
    _title_row(ws, "État des stocks", len(headers), 1)
    _col_headers(ws, headers, 2)

    for i, s in enumerate(stocks):
        valeur = round(s.quantite * (s.cout_unitaire or 0), 2)
        en_alerte = s.seuil_alerte > 0 and s.quantite <= s.seuil_alerte
        row = i + 3
        _data_row(ws, [
            s.nom,
            s.categorie.value if s.categorie else "—",
            round(s.quantite, 2),
            s.unite or "—",
            s.seuil_alerte,
            s.cout_unitaire,
            valeur,
            s.notes or "",
        ], row=row, alternate=i % 2 == 0)
        if en_alerte:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = _header_fill(AMBER_LIGHT)

    widths = [24, 14, 12, 10, 14, 18, 16, 24]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)


# ── Feuille Recommandations ───────────────────────────────────────────
def _sheet_recommandations(wb, recommandations):
    ws = wb.create_sheet("Recommandations")
    headers = ["Date", "Auteur", "Priorité", "Statut", "Contenu"]
    _title_row(ws, "Recommandations", len(headers), 1)
    _col_headers(ws, headers, 2)

    for i, r in enumerate(recommandations):
        row = i + 3
        _data_row(ws, [
            r.date.strftime("%d/%m/%Y") if r.date else "",
            r.auteur or "—",
            r.priorite.value if r.priorite else "—",
            r.statut.value if r.statut else "—",
            r.contenu,
        ], row=row, alternate=i % 2 == 0)
        if r.priorite and r.priorite == models.PrioriteEnum.haute:
            ws.cell(row=row, column=3).fill = _header_fill(RED_LIGHT)

    widths = [14, 18, 12, 14, 60]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)
    ws.column_dimensions["E"].alignment = Alignment(wrap_text=True)


# ── Endpoint ──────────────────────────────────────────────────────────
@router.get("/bilan/{ferme_id}")
def export_excel(
    ferme_id: int,
    annee: int = Query(default=2025),
    db: Session = Depends(get_db),
    user=Depends(get_current_user)
):
    ferme = db.query(models.Ferme).filter(models.Ferme.id == ferme_id).first()
    if not ferme:
        raise HTTPException(status_code=404, detail="Ferme introuvable")

    parcelles = db.query(models.Parcelle).filter(models.Parcelle.ferme_id == ferme_id).all()
    parcelle_ids = [p.id for p in parcelles]
    parcelles_map = {p.id: p for p in parcelles}

    recoltes = db.query(models.Recolte).filter(
        models.Recolte.parcelle_id.in_(parcelle_ids),
        extract("year", models.Recolte.date) == annee
    ).order_by(models.Recolte.date).all() if parcelle_ids else []

    traitements = db.query(models.Traitement).filter(
        models.Traitement.parcelle_id.in_(parcelle_ids),
        extract("year", models.Traitement.date) == annee
    ).order_by(models.Traitement.date).all() if parcelle_ids else []

    stocks = db.query(models.Stock).filter(models.Stock.ferme_id == ferme_id).all()
    stock_ids = [s.id for s in stocks]

    mouvements = db.query(models.MouvementStock).filter(
        models.MouvementStock.stock_id.in_(stock_ids),
        extract("year", models.MouvementStock.date) == annee
    ).all() if stock_ids else []

    recommandations = db.query(models.Recommandation).filter(
        models.Recommandation.ferme_id == ferme_id
    ).order_by(models.Recommandation.date.desc()).all()

    # Générer le fichier Excel
    wb = openpyxl.Workbook()
    _sheet_resume(wb, ferme, annee, parcelles, recoltes, mouvements, recommandations)
    _sheet_recoltes(wb, recoltes, parcelles_map, annee)
    _sheet_traitements(wb, traitements, parcelles_map, annee)
    _sheet_stocks(wb, stocks)
    _sheet_recommandations(wb, recommandations)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"bilan_{ferme.nom.replace(' ', '_')}_{annee}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
